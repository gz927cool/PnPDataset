import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

source_dir = r"c:\Users\001\Desktop\list\04-Enrich"
output_dir = r"c:\Users\001\Desktop\list\06-Excel"

def apply_styles(filename):
    wb = load_workbook(filename)
    ws = wb.active
    
    # Define styles
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    # Type colors
    type_colors = {
        'Institution': PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"), # Light Green
        'Person': PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),      # Light Yellow
        'Place': PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),       # Light Red
        'Event': PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid"),       # Light Blue
        'Subject': PatternFill(start_color="EDEDED", end_color="EDEDED", fill_type="solid")      # Gray
    }
    
    # Apply Header Style
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        
    # Find 'Type' column index
    type_col_idx = None
    for cell in ws[1]:
        if cell.value == 'Type':
            type_col_idx = cell.column
            break
            
    # Iterate rows
    for row in ws.iter_rows(min_row=2):
        # Apply Type color if column exists
        if type_col_idx:
            type_cell = row[type_col_idx - 1]
            type_val = type_cell.value
            if type_val in type_colors:
                type_cell.fill = type_colors[type_val]
        
        # Add borders to all cells
        thin_border = Border(left=Side(style='thin'), 
                             right=Side(style='thin'), 
                             top=Side(style='thin'), 
                             bottom=Side(style='thin'))
        for cell in row:
            cell.border = thin_border

    # Auto-adjust column widths
    for column_cells in ws.columns:
        length = max(len(str(cell.value) if cell.value else "") for cell in column_cells)
        # Cap width at 50 to avoid huge columns
        length = min(length, 50) 
        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = length + 2

    wb.save(filename)

def convert_files():
    csv_files = [f for f in os.listdir(source_dir) if f.endswith('.csv')]
    
    for filename in csv_files:
        print(f"Converting {filename}...")
        input_path = os.path.join(source_dir, filename)
        output_filename = filename.replace('.csv', '.xlsx')
        output_path = os.path.join(output_dir, output_filename)
        
        # Read CSV
        # Try reading with different encodings just in case
        try:
            df = pd.read_csv(input_path, encoding='utf-8')
        except UnicodeDecodeError:
            df = pd.read_csv(input_path, encoding='latin-1')
            
        # Write to Excel
        df.to_excel(output_path, index=False)
        
        # Apply Styles
        apply_styles(output_path)
        
    print(f"Done. Excel files saved to {output_dir}")

if __name__ == "__main__":
    convert_files()
